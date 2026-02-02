import openpyxl
book = openpyxl.load_workbook("C:\\Users\\yuvra\\OneDrive\\Documents\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2) # to read the excel value
print(cell.value)
sheet.cell(row=2,column=2).value = "Yuvraj"

print (sheet.cell(row=2,column=2).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)
print(sheet['D1'].value)

# PRINT ALL THE VALUES PRESENT IN EXCEL SHEET

for i in range(1,sheet.max_row+1): # to get rows
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(1,sheet.max_column+1): # to get columns
            print(sheet.cell(row=i,column=j).value)